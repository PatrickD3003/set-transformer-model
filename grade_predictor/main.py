#!/usr/bin/env python
# coding: utf-8

# In[ ]:


# import modules
import torch    


# In[ ]:


import numpy as np
import json
import copy
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader, Subset
from torch.nn.utils.rnn import pad_sequence
import os
import csv
from models.modules_modified import ISAB, SAB, PMA
import pandas as pd
from collections import defaultdict

from sklearn.model_selection import train_test_split
from sklearn.utils.class_weight import compute_class_weight


# In[ ]:


# import models
from models.ensemble import (
    SoftVotingEnsemble,
    GeometricMeanEnsemble,
    MedianEnsemble,
    TrimmedMeanEnsemble,
    StackingEnsemble,
    AdaBoostEnsemble,
    GBMEnsemble,
    XGBoostEnsemble,
    LightGBMEnsemble,
)

from models.classifier import (
    SetTransformerClassifierXY,
    SetTransformerClassifierXYAdditive,
    SetTransformerClassifier,
    DeepSetClassifierXYAdditive,
    DeepSetClassifierXY,
    DeepSetClassifier,
)

from models.ordinal import (
    SetTransformerOrdinalXY,
    SetTransformerOrdinalXYAdditive,
    SetTransformerOrdinal,
    DeepSetOrdinalXYAdditive,
    DeepSetOrdinalXY,
    DeepSetOrdinal,
    OrdinalSoftVotingEnsemble,
    OrdinalGeometricMeanEnsemble,
    OrdinalMedianEnsemble,
    OrdinalTrimmedMeanEnsemble,
    OrdinalStackingEnsemble,
    OrdinalGBMEnsemble,
    OrdinalXGBoostEnsemble,
    OrdinalLightGBMEnsemble,
    OrdinalAdaBoostEnsemble,
)

from models.utils_ordinal import ordinal_logistic_loss, cumulative_to_labels, threshold_accuracy


# In[ ]:


# Mappings --------------------------------------------------------
# Map each hold like "A1"…"K18" to an integer 0…(11*18−1)=197
cols = [chr(c) for c in range(ord('A'), ord('K')+1)]
rows = list(range(1, 19))
hold_to_idx = {f"{c}{r}": i for i, (c, r) in enumerate((c, r) for r in rows for c in cols)}


# Map grades "V4"…"V11" 
grade_to_label = {f"V{i}": i - 4 for i in range(4, 12)}  
label_to_grade = {v: k for k, v in grade_to_label.items()}
print(hold_to_idx)

# number of iterations 
num_iterations = 25


# In[ ]:


# Holds difficulty data --------------------------------------------------------
hold_difficulty = {}
with open("data/hold_difficulty.txt", "r") as f:
    for line in f:
        if ":" not in line:
            continue  # skip malformed line
        hold, rest = line.strip().split(":", 1)
        parts = rest.strip().split(",")
        difficulty = int(parts[0].strip())
        types = [t.strip() for t in parts[1:]]
        hold_difficulty[hold.strip()] = (difficulty, types)
    print("successfully parsed hold difficulty file")

# prepare type vocabulary
unique_types = set()
for _, (_, types) in hold_difficulty.items():
    unique_types.update(types)

type_to_idx = {t: i for i, t in enumerate(sorted(unique_types))}
print(f"successfully prepare type vocabulary")


# In[ ]:


# assign x,y position to each holds -------------------------------
import string

# Board columns A–K → indices 0–10
cols = list(string.ascii_uppercase[:11])  # A–K
# Rows 1–18 → indices 0–17
rows = list(range(1, 19))  # 1–18

# Generate hold_to_coord dictionary
hold_to_coord = {}

for x, col in enumerate(cols):
    for y, row in enumerate(rows):
        hold_name = f"{col}{row}"
        hold_to_coord[hold_name] = (x, y)

print("successfully created (x,y) position to each hold:")
print(hold_to_coord)


# In[ ]:


class MoonBoardDataset(Dataset):
    def __init__(self, json_path, hold_to_idx, grade_to_label, hold_difficulty, type_to_idx, hold_to_coord, max_difficulty=10):
        self.hold_to_idx = hold_to_idx
        self.grade_to_label = grade_to_label
        self.hold_difficulty = hold_difficulty
        self.type_to_idx = type_to_idx
        self.hold_to_coord = hold_to_coord
        self.max_difficulty = max_difficulty

        with open(json_path, 'r') as f:
            self.raw = [json.loads(line) for line in f]

    def __len__(self):
        return len(self.raw)

    def __getitem__(self, idx):
        item = self.raw[idx]
        holds = item['holds']

        hold_idxs = []
        diff_values = []
        type_vecs = []
        xy_coords = []

        for h in holds:
            hold_idxs.append(self.hold_to_idx[h])

            difficulty, types = self.hold_difficulty[h]
            diff_values.append(difficulty / self.max_difficulty)

            # multi-hot vector
            type_vec = torch.zeros(len(self.type_to_idx), dtype=torch.float)
            for t in types:
                if t in self.type_to_idx:
                    type_vec[self.type_to_idx[t]] = 1.0
            type_vecs.append(type_vec)

            # normalized (x, y)
            x, y = self.hold_to_coord[h]
            xy_coords.append(torch.tensor([x / 10.0, y / 17.0], dtype=torch.float))

        return {
            "indices": torch.tensor(hold_idxs, dtype=torch.long),
            "difficulty": torch.tensor(diff_values, dtype=torch.float),
            "type": torch.stack(type_vecs),       # (N, T)
            "xy": torch.stack(xy_coords)          # (N, 2)
        }, torch.tensor(self.grade_to_label[item['grade']], dtype=torch.long)


# In[ ]:


from torch.utils.data import WeightedRandomSampler

# --- Set Hyperparameters ---
json_path = './data/cleaned_moonboard2024_grouped.json'
embed_dim = 64
batch_size = 16
lr = 1e-4
epochs = 20

# Check for CUDA (NVIDIA), then MPS (Mac Silicon), then fallback to CPU
if torch.cuda.is_available():
    device = torch.device("cuda")
elif torch.backends.mps.is_available():
    device = torch.device("mps")
else:
    device = torch.device("cpu")

print(f"Using device: {device}")

XY_MODELS = {
    'set_transformer_xy',
    'set_transformer_additive',
    'deepset_xy',
    'deepset_xy_additive',
    'set_transformer_ordinal_xy',
    'set_transformer_ordinal_xy_additive',
    'deepset_ordinal_xy',
    'deepset_ordinal_xy_additive',
}

ORDINAL_MODELS = {
    'set_transformer_ordinal',
    'set_transformer_ordinal_xy',
    'set_transformer_ordinal_xy_additive',
    'deepset_ordinal',
    'deepset_ordinal_xy',
    'deepset_ordinal_xy_additive',
}

# --- Collate Function Factory ---
def make_collate_fn(model_type):
    def collate_fn(batch):
        X_indices = [x['indices'] for x, _ in batch]
        X_difficulty = [x['difficulty'] for x, _ in batch]
        X_type = [x['type'] for x, _ in batch]
        y_batch = [y for _, y in batch]

        X_indices = pad_sequence(X_indices, batch_first=True)
        X_difficulty = pad_sequence(X_difficulty, batch_first=True)
        X_type = pad_sequence(X_type, batch_first=True)
        y_tensor = torch.stack(y_batch)

        if model_type in XY_MODELS:
            X_xy = [x['xy'] for x, _ in batch]
            X_xy = pad_sequence(X_xy, batch_first=True)
            return (X_indices, X_difficulty, X_type, X_xy), y_tensor
        else:
            return (X_indices,), y_tensor
    return collate_fn

# --- Dataset Loader ---
def load_dataset(json_path, hold_to_idx, grade_to_label, hold_difficulty, type_to_idx, hold_to_coord):
    return MoonBoardDataset(json_path, hold_to_idx, grade_to_label, hold_difficulty, type_to_idx, hold_to_coord)

# --- DataLoader Preparation ---
def prepare_dataloaders(dataset, grade_to_label, batch_size, collate_fn):
    targets = [grade_to_label[item['grade']] for item in dataset.raw]
    class_weights = compute_class_weight(class_weight='balanced', classes=np.unique(targets), y=targets)
    class_weights = torch.tensor(class_weights, dtype=torch.float).to(device)

    train_idx, val_idx = train_test_split(
        list(range(len(dataset))), test_size=0.2, stratify=targets, random_state=42
    )

    train_data = Subset(dataset, train_idx)
    val_data = Subset(dataset, val_idx)

    train_loader = DataLoader(train_data, batch_size=batch_size, shuffle=True, collate_fn=collate_fn)
    val_loader = DataLoader(val_data, batch_size=batch_size, shuffle=False, collate_fn=collate_fn)
    return train_loader, val_loader, class_weights, train_idx, val_idx

# --- Training ---
def train_model(model, train_loader, val_loader, criterion, optimizer, epochs, is_ordinal=False):
    for epoch in range(1, epochs + 1):
        model.train()
        total_loss = 0
        for X, y in train_loader:
            inputs = tuple(x.to(device) for x in X)
            y = y.to(device)
            payload = inputs[0] if len(inputs) == 1 else inputs
            outputs = model(payload)
            if is_ordinal:
                probs, logits = outputs
                loss = criterion(logits, y)
            else:
                logits = outputs
                loss = criterion(logits, y)
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            total_loss += loss.item()
        print(f"Epoch {epoch:02d} — loss: {total_loss / len(train_loader):.4f}")
    return model

# --- Main Per Model ---
def main(model_type):
    dataset = load_dataset(json_path, hold_to_idx, grade_to_label, hold_difficulty, type_to_idx, hold_to_coord)
    targets = [grade_to_label[item['grade']] for item in dataset.raw]
    num_classes = len(np.unique(targets))
    vocab_size = len(hold_to_idx)
    type_vec_dim = len(type_to_idx)
    is_ordinal = model_type in ORDINAL_MODELS

    if model_type == 'set_transformer':
        ModelClass = SetTransformerClassifier
        kwargs = dict(vocab_size=vocab_size, dim_in=embed_dim, num_classes=num_classes)
    elif model_type == 'set_transformer_xy':
        ModelClass = SetTransformerClassifierXY
        kwargs = dict(vocab_size=vocab_size, dim_in=embed_dim, num_classes=num_classes, type_vec_dim=type_vec_dim)
    elif model_type == 'set_transformer_additive':
        ModelClass = SetTransformerClassifierXYAdditive
        kwargs = dict(vocab_size=vocab_size, feat_dim=embed_dim, num_classes=num_classes, type_vec_dim=type_vec_dim)
    elif model_type == 'deepset':
        ModelClass = DeepSetClassifier
        kwargs = dict(vocab_size=vocab_size, dim_in=embed_dim, num_classes=num_classes)
    elif model_type == 'deepset_xy':
        ModelClass = DeepSetClassifierXY
        kwargs = dict(vocab_size=vocab_size, dim_in=embed_dim, num_classes=num_classes, type_vec_dim=type_vec_dim)
    elif model_type == 'deepset_xy_additive':
        ModelClass = DeepSetClassifierXYAdditive
        kwargs = dict(vocab_size=vocab_size, feat_dim=embed_dim, num_classes=num_classes, type_vec_dim=type_vec_dim)
    elif model_type == 'set_transformer_ordinal':
        ModelClass = SetTransformerOrdinal
        kwargs = dict(vocab_size=vocab_size, dim_in=embed_dim, num_classes=num_classes)
    elif model_type == 'set_transformer_ordinal_xy':
        ModelClass = SetTransformerOrdinalXY
        kwargs = dict(vocab_size=vocab_size, dim_in=embed_dim, num_classes=num_classes, type_vec_dim=type_vec_dim)
    elif model_type == 'set_transformer_ordinal_xy_additive':
        ModelClass = SetTransformerOrdinalXYAdditive
        kwargs = dict(vocab_size=vocab_size, feat_dim=embed_dim, num_classes=num_classes, type_vec_dim=type_vec_dim)
    elif model_type == 'deepset_ordinal':
        ModelClass = DeepSetOrdinal
        kwargs = dict(vocab_size=vocab_size, dim_in=embed_dim, num_classes=num_classes)
    elif model_type == 'deepset_ordinal_xy':
        ModelClass = DeepSetOrdinalXY
        kwargs = dict(vocab_size=vocab_size, dim_in=embed_dim, num_classes=num_classes, type_vec_dim=type_vec_dim)
    elif model_type == 'deepset_ordinal_xy_additive':
        ModelClass = DeepSetOrdinalXYAdditive
        kwargs = dict(vocab_size=vocab_size, feat_dim=embed_dim, num_classes=num_classes, type_vec_dim=type_vec_dim)
    else:
        raise ValueError(f"Unknown model_type: {model_type}")

    collate_fn = make_collate_fn(model_type)
    train_loader, val_loader, class_weights, train_idx, val_idx = prepare_dataloaders(dataset, grade_to_label, batch_size, collate_fn)

    model = ModelClass(**kwargs).to(device)
    model.is_ordinal = is_ordinal
    model.num_classes = num_classes

    if is_ordinal:
        def criterion_fn(logits, targets):
            return ordinal_logistic_loss(logits, targets)
    else:
        criterion_fn = torch.nn.CrossEntropyLoss(weight=class_weights)

    optimizer = torch.optim.Adam(model.parameters(), lr=lr)

    model = train_model(model, train_loader, val_loader, criterion_fn, optimizer, epochs, is_ordinal=is_ordinal)
    return train_loader, val_loader, model, dataset, train_idx, val_idx


def train_boosting_main(model_types, num_stages=5, weak_epochs=3):
    """
    Replaces the `main(mtype)` call for boosting models.
    Trains a sequential AdaBoost-style ensemble that can rotate through
    multiple base model types.
    
    Returns the same tuple as `main()`:
    (train_loader, val_loader, final_model, dataset, train_idx, val_idx)
    """
    if isinstance(model_types, str):
        model_types = [model_types]
    if not model_types:
        raise ValueError("At least one base model type is required for boosting.")

    print(f"===== Training Boosting Ensemble ({model_types}, {num_stages} stages) =====")
    
    # --- 1. Standard Dataset Setup (copied from `main`) ---
    dataset = load_dataset(json_path, hold_to_idx, grade_to_label, hold_difficulty, type_to_idx, hold_to_coord)
    targets = [grade_to_label[item['grade']] for item in dataset.raw]
    num_classes = len(np.unique(targets))
    vocab_size = len(hold_to_idx)
    type_vec_dim = len(type_to_idx)
    is_ordinal = False # Boosting classifiers, not ordinal models

    # Collate functions for every participating base model
    collate_fns = {mtype: make_collate_fn(mtype) for mtype in set(model_types)}

    # Get train/val split (we need the indices)
    train_idx, val_idx = train_test_split(
        list(range(len(dataset))), test_size=0.2, stratify=targets, random_state=42
    )
    
    train_data_subset = Subset(dataset, train_idx)
    val_data_subset = Subset(dataset, val_idx)
    
    # Per-model loaders (no shuffle) for evaluating each weak learner
    train_eval_loaders = {
        mtype: DataLoader(train_data_subset, batch_size=batch_size, shuffle=False, collate_fn=collate_fns[mtype])
        for mtype in collate_fns
    }
    val_loaders = {
        mtype: DataLoader(val_data_subset, batch_size=batch_size, shuffle=False, collate_fn=collate_fns[mtype])
        for mtype in collate_fns
    }

    # Use a superset collate function for the final ensemble (needs XY features if any member does)
    eval_collate_type = next((m for m in model_types if m in XY_MODELS), model_types[0])
    final_train_loader = train_eval_loaders[eval_collate_type]
    final_val_loader = val_loaders[eval_collate_type]

    # --- 2. Boosting-Specific Setup ---
    num_train_samples = len(train_idx)
    # Initialize uniform sample weights
    sample_weights = torch.full((num_train_samples,), 1.0 / num_train_samples, device=device)
    
    trained_models_list = [] # (name, model)
    model_alphas = []        # [alpha]
    stage_model_types = []   # Track which model was used at each stage
    
    # --- 3. The Sequential Training Loop ---
    for m in range(num_stages):
        stage_model_type = model_types[m % len(model_types)]
        print(f"--- Boosting Stage {m+1}/{num_stages} ({stage_model_type}) ---")
        
        # a. Create a new dataloader for this stage
        #    It samples from train_data_subset based on the *current* sample_weights
        sampler = WeightedRandomSampler(sample_weights.cpu(), num_train_samples, replacement=True)
        train_loader_stage = DataLoader(
            train_data_subset,
            batch_size=batch_size,
            sampler=sampler,
            collate_fn=collate_fns[stage_model_type],
        )

        # b. Create and train the weak learner
        #    We re-use the logic from your `main()` function to get the model
        if stage_model_type == 'set_transformer':
            ModelClass = SetTransformerClassifier
            kwargs = dict(vocab_size=vocab_size, dim_in=embed_dim, num_classes=num_classes)
        elif stage_model_type == 'set_transformer_xy':
            ModelClass = SetTransformerClassifierXY
            kwargs = dict(vocab_size=vocab_size, dim_in=embed_dim, num_classes=num_classes, type_vec_dim=type_vec_dim)
        elif stage_model_type == 'set_transformer_additive':
            ModelClass = SetTransformerClassifierXYAdditive
            kwargs = dict(vocab_size=vocab_size, feat_dim=embed_dim, num_classes=num_classes, type_vec_dim=type_vec_dim)
        elif stage_model_type == 'deepset':
            ModelClass = DeepSetClassifier
            kwargs = dict(vocab_size=vocab_size, dim_in=embed_dim, num_classes=num_classes)
        elif stage_model_type == 'deepset_xy':
            ModelClass = DeepSetClassifierXY
            kwargs = dict(vocab_size=vocab_size, dim_in=embed_dim, num_classes=num_classes, type_vec_dim=type_vec_dim)
        elif stage_model_type == 'deepset_xy_additive':
            ModelClass = DeepSetClassifierXYAdditive
            kwargs = dict(vocab_size=vocab_size, feat_dim=embed_dim, num_classes=num_classes, type_vec_dim=type_vec_dim)
        else:
            # This is the correct error message
            raise ValueError(f"Unsupported weak learner type for boosting: {stage_model_type}")

        model_m = ModelClass(**kwargs).to(device)
        
        # We can use standard CrossEntropyLoss because the *sampler* already weighted the data
        criterion_fn = torch.nn.CrossEntropyLoss()
        optimizer = torch.optim.Adam(model_m.parameters(), lr=lr)
        
        # c. Train this weak learner (RE-USING YOUR EXISTING `train_model` FUNCTION!)
        print(f"Training weak learner {m+1} for {weak_epochs} epochs...")
        model_m = train_model(
            model_m,
            train_loader_stage,
            val_loaders[stage_model_type],
            criterion_fn,
            optimizer,
            weak_epochs,
            is_ordinal=is_ordinal,
        )
        
        # d. Evaluate on *all* training data (unshuffled)
        model_m.eval()
        all_preds = []
        all_targets = []
        with torch.no_grad():
            for X, y in train_eval_loaders[stage_model_type]: # Use NON-shuffled loader
                inputs = tuple(x.to(device) for x in X)
                payload = inputs[0] if len(inputs) == 1 else inputs
                logits = model_m(payload)
                all_preds.append(logits.argmax(dim=1))
                all_targets.append(y.to(device))
        
        all_preds = torch.cat(all_preds)
        all_targets = torch.cat(all_targets)
        
        # e. Compute weighted error
        is_incorrect = (all_preds != all_targets).float() # [num_train_samples]
        err_m = (is_incorrect * sample_weights).sum() # Sum of weights of incorrect samples
        
        if err_m <= 0 or err_m >= (1.0 - 1.0 / num_classes):
            print(f"Stage {m+1} model is perfect or too weak (err={err_m:.4f}). Stopping.")
            if err_m <= 0: # Add perfect model and break
                model_alphas.append(1.0) # Use a reasonable weight
                trained_models_list.append((f"{stage_model_type}_boost_model_{m}", model_m))
                stage_model_types.append(stage_model_type)
            break
        
        # f. Compute model weight (alpha)
        alpha_m = torch.log((1.0 - err_m) / err_m) + torch.log(torch.tensor(num_classes - 1.0, device=device))
        
        # g. Update sample weights
        sample_weights *= torch.exp(alpha_m * is_incorrect)
        sample_weights /= sample_weights.sum() # Normalize
        
        # h. Save
        stage_model_types.append(stage_model_type)
        trained_models_list.append((f"{stage_model_type}_boost_model_{m}", model_m))
        model_alphas.append(alpha_m.item())
        print(f"Stage {m+1}: Error={err_m:.4f}, Alpha={alpha_m:.4f}")
    
    # --- 4. Build the Final Ensemble Model ---
    if not trained_models_list:
        raise RuntimeError("Boosting training failed, no models were trained.")
        
    print(f"Building final ensemble with {len(trained_models_list)} models.")
    
    # *** USE THE NEW NAME HERE ***
    final_ensemble = AdaBoostEnsemble(trained_models_list, weights=model_alphas, freeze_members=True).to(device)
    
    final_ensemble.is_ordinal = is_ordinal
    final_ensemble.num_classes = num_classes
    final_ensemble.stage_model_types = stage_model_types
    
    # Return the same "package" as main()
    # `train_eval_loader` is the unshuffled train loader, which `compare_models`
    # can use for ensemble training (like stacking) if needed.
    return final_train_loader, final_val_loader, final_ensemble, dataset, train_idx, val_idx


# In[ ]:


import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import confusion_matrix
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

BASE_MODEL_TYPES = [
    "set_transformer",
    "deepset",
    "set_transformer_xy",
    "deepset_xy",
    "set_transformer_additive",
    "deepset_xy_additive"
]

BOOSTING_TYPES = {
    "adaboost_all": list(BASE_MODEL_TYPES),
    "adaboost_set_transformer": ["set_transformer", "set_transformer_xy", "set_transformer_additive"],
    "adaboost_deepset": ["deepset", "deepset_xy", "deepset_xy_additive"],
}

MODEL_TYPES = BASE_MODEL_TYPES + list(BOOSTING_TYPES.keys())

ENSEMBLE_TYPES = [
    "soft_voting_ensemble",
    "stacking_ensemble",
    "gbm_ensemble",
    "xgboost_ensemble",
    # "lightgbm_ensemble"
    # "geometric_mean_ensemble",
    # "median_ensemble",
    # "trimmed_mean_ensemble",
]

MODEL_COUNT_COLUMNS = {name: f"{name}_count" for name in MODEL_TYPES + ENSEMBLE_TYPES}


# --- plot confusion matrix and save to excel---
def save_confusion_matrix_to_excel(y_true, y_pred, class_labels, model_name, excel_path):
    # Plot confusion matrix and save as image
    cm = confusion_matrix(y_true, y_pred, labels=range(len(class_labels)), normalize='true')
    plt.figure(figsize=(8, 6))
    sns.heatmap(cm, annot=True, fmt=".2f", cmap="Blues", xticklabels=class_labels, yticklabels=class_labels)
    plt.title(f"Confusion Matrix: {model_name}")
    plt.xlabel("Predicted Grade")
    plt.ylabel("Actual Grade")
    plt.tight_layout()
    img_path = f"result/confusion_{model_name}.png"
    plt.savefig(img_path)
    plt.close()

    # Insert image into Excel (new sheet per model)
    wb = load_workbook(excel_path)
    if model_name in wb.sheetnames:
        ws = wb[model_name]
    else:
        ws = wb.create_sheet(title=model_name)
    img = XLImage(img_path)
    ws.add_image(img, "A1")
    wb.save(excel_path)
    print(f"Confusion matrix for {model_name} saved and inserted into {excel_path} (sheet: {model_name})")


# --- export the predictions to excel ---
def _update_outlier_excel(df_all_preds, model_name, outlier_filename="result/outlier.xlsx", sheet_name="outliers", threshold=3):
    """
    From a DataFrame with columns [problem_name, y_true, y_pred, diff],
    keep rows where abs(diff) > threshold and aggregate per problem_name:
        - count = number of times flagged
        - per-model counts = number of times flagged per model across runs
        - y_true = mode (most frequent true label)
        - y_pred_avg = average predicted label across occurrences
    Save to outlier.xlsx.
    """
    current_model_col = MODEL_COUNT_COLUMNS.get(model_name, f"{model_name}_count")
    model_count_columns = dict(MODEL_COUNT_COLUMNS)
    if model_name not in MODEL_COUNT_COLUMNS:
        model_count_columns[model_name] = current_model_col

    # Filter outliers
    outliers = df_all_preds.loc[df_all_preds["diff"].abs() > threshold,
                                ["problem_name", "y_true", "y_pred"]]
    if outliers.empty:
        print(f"No outliers (abs(diff) > {threshold}). Skipped creating outlier.xlsx.")
        return

    # Group & aggregate
    grouped = (outliers
               .groupby("problem_name")
               .agg(
                   count=("problem_name", "size"),
                   y_true=("y_true", lambda x: x.mode().iat[0] if not x.mode().empty else x.iloc[0]),
                   y_pred_avg=("y_pred", lambda x: round(pd.to_numeric(x, errors="coerce").mean(), 2))
               )
               .reset_index())

    for col in model_count_columns.values():
        if col not in grouped.columns:
            grouped[col] = 0
    grouped[current_model_col] = grouped["count"]

    # If a previous file exists, merge and accumulate counts
    if os.path.exists(outlier_filename):
        try:
            existing = pd.read_excel(outlier_filename, sheet_name=sheet_name)
            for col in model_count_columns.values():
                if col not in existing.columns:
                    existing[col] = 0
            if set(existing.columns) >= {"problem_name", "count", "y_true", "y_pred_avg"}:
                merged = pd.concat([existing, grouped], ignore_index=True, sort=False)
                for col in model_count_columns.values():
                    if col not in merged.columns:
                        merged[col] = 0
                agg_map = {
                    "count": ("count", "sum"),
                    "y_true": ("y_true", lambda x: x.mode().iat[0] if not x.mode().empty else x.iloc[0]),
                    "y_pred_avg": ("y_pred_avg", "mean")
                }
                agg_map.update({col: (col, "sum") for col in model_count_columns.values()})
                grouped = (merged
                           .groupby("problem_name")
                           .agg(**agg_map)
                           .reset_index())
            # else keep grouped as new
        except Exception:
            pass

    grouped["y_pred_avg"] = pd.to_numeric(grouped["y_pred_avg"], errors="coerce").round(2)
    grouped["count"] = grouped["count"].fillna(0).astype(int)
    for col in model_count_columns.values():
        if col in grouped.columns:
            grouped[col] = grouped[col].fillna(0).astype(int)

    ordered_cols = ["problem_name", "count"]
    ordered_cols.extend([model_count_columns[name] for name in MODEL_TYPES + ENSEMBLE_TYPES if model_count_columns[name] in grouped.columns])
    if current_model_col in grouped.columns and current_model_col not in ordered_cols:
        ordered_cols.append(current_model_col)
    ordered_cols.extend(["y_true", "y_pred_avg"])
    ordered_cols.extend([col for col in grouped.columns if col not in ordered_cols])
    grouped = grouped[ordered_cols]

    # Save
    with pd.ExcelWriter(outlier_filename, engine="openpyxl", mode="w") as writer:
        grouped.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Outliers saved to: {os.path.abspath(outlier_filename)}")


def export_predictions_to_excel(model, dataloader, device, grade_to_label, excel_path, sheet_name, model_name=None):
    results = []
    raw_dataset = dataloader.dataset.dataset  # MoonBoardDataset
    indices = dataloader.dataset.indices      # Subset indices
    label_to_grade = {v: k for k, v in grade_to_label.items()}
    current_index = 0

    model.eval()
    with torch.no_grad():
        for X, y in dataloader:
            if isinstance(X, tuple):
                inputs = tuple(x.to(device) for x in X)
                payload = inputs[0] if len(inputs) == 1 else inputs
            else:
                payload = X.to(device)
            outputs = model(payload)
            if isinstance(outputs, tuple):
                if getattr(model, "is_ordinal", False):
                    probs = outputs[0]
                    preds_tensor = cumulative_to_labels(probs)
                else:
                    probs = outputs[0]
                    preds_tensor = probs.argmax(dim=1)
            else:
                preds_tensor = outputs.argmax(dim=1)
            y = y.to(device)
            preds_cpu = preds_tensor.detach().cpu()
            y_cpu = y.detach().cpu()
            for i in range(y_cpu.size(0)):
                real_label = int(y_cpu[i].item())
                pred_label = int(preds_cpu[i].item())
                dataset_index = indices[current_index]
                current_index += 1
                raw_item = raw_dataset.raw[dataset_index]
                problem_name = raw_item.get('problem_name', f"problem_{dataset_index}")
                results.append({
                    "problem_name": problem_name,
                    "y_true": real_label,  # keep numeric for averaging/aggregation
                    "y_pred": pred_label,
                    "diff": real_label - pred_label
                })

    df = pd.DataFrame(results)

    if model_name is None:
        model_name = sheet_name
    df["model_name"] = model_name
    df["model"] = model_name

    # 1) Save all predictions into your main Excel file
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        # Convert numeric labels back to grade strings for readability
        df_out = df.copy()
        df_out["y_true"] = df_out["y_true"].map(lambda x: label_to_grade.get(x, f"Unknown({x})"))
        df_out["y_pred"] = df_out["y_pred"].map(lambda x: label_to_grade.get(x, f"Unknown({x})"))
        df_out.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"Predictions for {sheet_name} exported to: {excel_path}")

    # 2) Create/update outlier.xlsx (problem_name, count, per-model counts, y_true, y_pred_avg)
    _update_outlier_excel(df, model_name=model_name, outlier_filename="result/outlier.xlsx", sheet_name="outliers", threshold=3)


# --- compute training and validation accuracy ---
def compute_accuracy(model, dataloader, device):
    strict_correct, loose_correct, total = 0, 0, 0
    y_true, y_pred = [], []
    model.eval()
    with torch.no_grad():
        for X, y in dataloader:
            X = tuple(x.to(device) for x in X)
            y = y.to(device)

            payload = X[0] if len(X) == 1 else X
            outputs = model(payload)

            if isinstance(outputs, tuple):
                if getattr(model, "is_ordinal", False):
                    probs = outputs[0]
                    preds_tensor = cumulative_to_labels(probs)
                else:
                    probs = outputs[0]
                    preds_tensor = probs.argmax(dim=1)
            else:
                preds_tensor = outputs.argmax(dim=1)

            if isinstance(preds_tensor, torch.Tensor):
                preds_tensor = preds_tensor.to(y.device)
            preds = preds_tensor
            total += y.size(0)
            strict_correct += (preds == y).sum().item()
            loose_correct += ((preds - y).abs() <= 1).sum().item()
            y_true.extend(y.cpu().numpy())
            y_pred.extend(preds.detach().cpu().numpy())
    strict_acc = 100.0 * strict_correct / total
    loose_acc = 100.0 * loose_correct / total
    return strict_acc, loose_acc, y_true, y_pred


def log_accuracy_to_csv(model_type, train_strict_acc, train_loose_acc, val_strict_acc, val_loose_acc, csv_path="result/accuracy.csv"):
    file_exists = os.path.isfile(csv_path)
    with open(csv_path, mode='a', newline='') as csvfile:
        writer = csv.writer(csvfile)
        if not file_exists:
            writer.writerow([
                "model",
                "Train Strict Accuracy (%)",
                "Train ±1 Grade Accuracy (%)",
                "Val Strict Accuracy (%)",
                "Val ±1 Grade Accuracy (%)"
            ])
        writer.writerow([
            model_type,
            round(train_strict_acc, 2),
            round(train_loose_acc, 2),
            round(val_strict_acc, 2),
            round(val_loose_acc, 2)
        ])


# In[ ]:


def train_stacking_meta_model(stacking_model, dataloader, device, epochs=5, lr=1e-3):
    """Train stacking meta-learner on frozen base model outputs."""
    if epochs <= 0:
        return
    stacking_model.meta_model.train()
    for member in stacking_model.models.values():
        member.eval()
    optimizer = torch.optim.Adam(stacking_model.meta_model.parameters(), lr=lr)
    criterion = nn.CrossEntropyLoss()

    for epoch in range(epochs):
        epoch_loss = 0.0
        num_samples = 0
        for X, y in dataloader:
            inputs = tuple(x.to(device) for x in X)
            targets = y.to(device)
            member_feats = stacking_model._member_features(inputs)
            M, B, F = member_feats.shape
            if stacking_model.combine == "mean":
                feat = member_feats.mean(dim=0)
            else:
                feat = member_feats.permute(1, 0, 2).reshape(B, M * F)

            logits = stacking_model.meta_model(feat)
            loss = criterion(logits, targets)

            optimizer.zero_grad()
            loss.backward()
            optimizer.step()

            batch_size = targets.size(0)
            epoch_loss += loss.item() * batch_size
            num_samples += batch_size

        denom = num_samples if num_samples > 0 else 1
        avg_loss = epoch_loss / denom
        print(f"Stacking meta epoch {epoch + 1}: loss={avg_loss:.4f}")

    stacking_model.meta_model.eval()


def train_tree_meta_model(tree_ensemble, dataloader, device):
    """Fit a tree-based meta-learner (GBM/XGBoost/LightGBM) on frozen base outputs."""
    tree_ensemble.eval()
    for member in tree_ensemble.models.values():
        member.eval()
    feature_blocks = []
    target_blocks = []
    for X, y in dataloader:
        inputs = tuple(x.to(device) for x in X)
        member_feats = tree_ensemble._member_features(inputs)
        feat = tree_ensemble._build_feature_matrix(member_feats)
        feature_blocks.append(feat.detach().cpu().numpy())
        target_blocks.append(y.detach().cpu().numpy())
    if not feature_blocks:
        raise RuntimeError("No data available to fit tree-based meta learner.")
    features = np.concatenate(feature_blocks, axis=0)
    targets = np.concatenate(target_blocks, axis=0)
    tree_ensemble.fit_meta_model(features, targets)

def infer_stacking_feature_dim(stacking_model, dataloader, device):
    """
    Determine the flattened feature dimension seen by the stacking meta-model.
    """
    stacking_model.eval()
    with torch.no_grad():
        for X, _ in dataloader:
            inputs = tuple(x.to(device) for x in X)
            member_feats = stacking_model._member_features(inputs)
            M, _, F = member_feats.shape
            return F if stacking_model.combine == "mean" else M * F
    raise RuntimeError("Unable to infer stacking feature dimension (empty dataloader?).")


def build_ensemble_models(
    ensemble_names,
    base_model_items,
    ensemble_weights,
    num_classes,
    device,
    train_loader,
    stacking_meta_epochs=5,
    stacking_meta_lr=1e-3,
    label_suffix="",
):
    """Create configured ensemble models from trained base models."""
    ensembles = {}
    base_items = list(base_model_items)

    if not base_items:
        return ensembles

    def _resolve_group_weights(items):
        if ensemble_weights is None:
            return None
        if isinstance(ensemble_weights, dict):
            filtered = {name: ensemble_weights[name] for name, _ in items if name in ensemble_weights}
            if len(filtered) != len(items):
                missing = [name for name, _ in items if name not in filtered]
                if missing:
                    print(f"Warning: missing weights for {missing}; using uniform weights.")
                return None
            return filtered
        weight_list = list(ensemble_weights)
        if len(weight_list) != len(items):
            print("Warning: weight list length mismatch; using uniform weights.")
            return None
        return weight_list

    resolved_weights = _resolve_group_weights(base_items)

    for base_name in ensemble_names:
        cloned_items = [(model_name, copy.deepcopy(model)) for model_name, model in base_items]
        if not cloned_items:
            continue

        weights = resolved_weights
        if isinstance(weights, list):
            weights = list(weights)

        ensemble_key = f"{base_name}{label_suffix}" if label_suffix else base_name

        if base_name == "soft_voting_ensemble":
            ensemble_model = SoftVotingEnsemble(cloned_items, weights=weights, freeze_members=True).to(device)
        elif base_name == "geometric_mean_ensemble":
            ensemble_model = GeometricMeanEnsemble(cloned_items, weights=weights, freeze_members=True).to(device)
        elif base_name == "median_ensemble":
            ensemble_model = MedianEnsemble(cloned_items, weights=weights, freeze_members=True).to(device)
        elif base_name == "trimmed_mean_ensemble":
            ensemble_model = TrimmedMeanEnsemble(cloned_items, weights=weights, freeze_members=True, trim_frac=0.2).to(device)
        elif base_name == "stacking_ensemble":
            placeholder_dim = len(cloned_items) * num_classes
            meta_model = nn.Linear(placeholder_dim, num_classes).to(device)
            ensemble_model = StackingEnsemble(
                cloned_items,
                weights=weights,
                freeze_members=True,
                meta_model=meta_model,
                feature_source="logits+internal",
                combine="concat",
            ).to(device)
            inferred_dim = infer_stacking_feature_dim(ensemble_model, train_loader, device)
            if inferred_dim != placeholder_dim:
                ensemble_model.meta_model = nn.Linear(inferred_dim, num_classes).to(device)
            train_stacking_meta_model(
                ensemble_model,
                train_loader,
                device,
                epochs=stacking_meta_epochs,
                lr=stacking_meta_lr,
            )
        elif base_name == "gbm_ensemble":
            try:
                ensemble_model = GBMEnsemble(
                    cloned_items,
                    weights=weights,
                    freeze_members=True,
                    num_classes=num_classes,
                    feature_source="logits",
                    combine="concat",
                    meta_kwargs={"random_state": 42},
                ).to(device)
            except ImportError as exc:
                print(f"Skipping {base_name}: {exc}")
                continue
            train_tree_meta_model(ensemble_model, train_loader, device)
        elif base_name == "xgboost_ensemble":
            try:
                ensemble_model = XGBoostEnsemble(
                    cloned_items,
                    weights=weights,
                    freeze_members=True,
                    num_classes=num_classes,
                    feature_source="logits",
                    combine="concat",
                    meta_kwargs={"n_estimators": 300, "learning_rate": 0.05, "max_depth": 4},
                ).to(device)
            except ImportError as exc:
                print(f"Skipping {base_name}: {exc}")
                continue
            train_tree_meta_model(ensemble_model, train_loader, device)
        elif base_name == "lightgbm_ensemble":
            try:
                ensemble_model = LightGBMEnsemble(
                    cloned_items,
                    weights=weights,
                    freeze_members=True,
                    num_classes=num_classes,
                    feature_source="logits",
                    combine="concat",
                    meta_kwargs={"n_estimators": 300, "learning_rate": 0.05, "max_depth": -1},
                ).to(device)
            except ImportError as exc:
                print(f"Skipping {base_name}: {exc}")
                continue
            train_tree_meta_model(ensemble_model, train_loader, device)
        else:
            print(f"Unknown ensemble type '{base_name}', skipping.")
            continue

        ensemble_model.eval()
        ensembles[ensemble_key] = ensemble_model

    return ensembles


# In[ ]:


def compare_models(
    model_types=None,
    include_ensemble=True,
    ensemble_weights=None,
    ensemble_types=None,
    stacking_meta_epochs=5,
    stacking_meta_lr=1e-3,
    boosting_num_stages=5, 
    boosting_weak_epochs=3 
):
    model_types = model_types or MODEL_TYPES
    results = []
    excel_path = "result/model_comparison_results.xlsx"
    class_labels = [f"V{i}" for i in range(4, 12)]

    # BOOSTING_TYPES (defined globally) maps each boosting variant to its base learner.

    trained_models = {}
    base_dataset = None
    base_train_idx = None
    base_val_idx = None
    num_classes = None

    # --- THIS IS THE MODIFIED LOOP ---
    for idx, mtype in enumerate(model_types):
        print(f"===== Processing {mtype} =====")
        
        if mtype in BOOSTING_TYPES:
            # --- This is a SEQUENTIAL (Boosting) Model ---
            base_model_types = BOOSTING_TYPES[mtype]
            # Call our new boosting trainer
            train_loader, val_loader, model, dataset, train_idx, val_idx = train_boosting_main(
                model_types=base_model_types,
                num_stages=boosting_num_stages,
                weak_epochs=boosting_weak_epochs
            )
        else:
            # --- This is a PARALLEL (Standard) Model ---
            print(f"===== Training {mtype} =====")
            train_loader, val_loader, model, dataset, train_idx, val_idx = main(mtype)

        # --- THE REST OF THE LOOP IS UNCHANGED ---
        # Because both `main` and `train_boosting_main` return the same
        # tuple, the evaluation code works perfectly for both.
        model.eval()
        trained_models[mtype] = model
        if base_dataset is None:
            base_dataset = dataset
            base_train_idx = train_idx
            base_val_idx = val_idx
        if num_classes is None:
            num_classes = getattr(model, "num_classes", None)

        train_strict_acc, train_loose_acc, _, _ = compute_accuracy(model, train_loader, device)
        val_strict_acc, val_loose_acc, y_true, y_pred = compute_accuracy(model, val_loader, device)

        log_accuracy_to_csv(mtype, train_strict_acc, train_loose_acc, val_strict_acc, val_loose_acc)

        results.append({
            "Model Type": mtype,
            "Train Strict Accuracy (%)": train_strict_acc,
            "Train ±1 Grade Accuracy (%)": train_loose_acc,
            "Val Strict Accuracy (%)": val_strict_acc,
            "Val ±1 Grade Accuracy (%)": val_loose_acc,
        })

        if idx == 0:
            df_results = pd.DataFrame(results)
            df_results.to_excel(excel_path, index=False)

        # save_confusion_matrix_to_excel(y_true, y_pred, class_labels, mtype, excel_path)
        export_predictions_to_excel(
            model,
            val_loader,
            device,
            grade_to_label,
            excel_path,
            sheet_name=f"{mtype}_preds",
            model_name=mtype,
        )

    # --- ENSEMBLE EVALUATION SECTION ---
    if include_ensemble and trained_models:
        effective_ensemble_types = ensemble_types or ENSEMBLE_TYPES
        if not effective_ensemble_types:
            print("No ensemble types specified; skipping ensemble evaluation.")
        else:
            if base_dataset is None or base_train_idx is None or base_val_idx is None:
                raise RuntimeError("Dataset indices are unavailable for ensemble evaluation.")
            if num_classes is None:
                raise RuntimeError("Unable to determine number of classes for ensembles.")

            print("===== Evaluating ensembles =====")
            collate_fn = make_collate_fn("set_transformer_xy")
            train_subset = Subset(base_dataset, base_train_idx)
            val_subset = Subset(base_dataset, base_val_idx)
            ensemble_train_loader = DataLoader(train_subset, batch_size=batch_size, shuffle=False, collate_fn=collate_fn)
            ensemble_val_loader = DataLoader(val_subset, batch_size=batch_size, shuffle=False, collate_fn=collate_fn)

            # Filter trained_models to EXCLUDE boosting models from being ensembled
            base_items = [(name, model) for name, model in trained_models.items() if name not in BOOSTING_TYPES]

            # Now ensemble_groups will only contain the original base models
            ensemble_groups = [
                ("all", base_items),
                ("set_transformer", [(name, model) for name, model in base_items if "set_transformer" in name]),
                ("deepset", [(name, model) for name, model in base_items if "deepset" in name]),
            ]

            for group_name, group_items in ensemble_groups:
                if not group_items:
                    print(f"Skipping {group_name} ensemble group (no models).")
                    continue

                print(f"--- Evaluating {group_name} ensembles ({len(group_items)} models) ---")
                ensembles = build_ensemble_models(
                    effective_ensemble_types,
                    group_items,
                    ensemble_weights,
                    num_classes,
                    device,
                    ensemble_train_loader,
                    stacking_meta_epochs=stacking_meta_epochs,
                    stacking_meta_lr=stacking_meta_lr,
                    label_suffix=f"_{group_name}",
                )

                if not ensembles:
                    print(f"No ensembles constructed for group {group_name}.")
                    continue

                for name, ensemble_model in ensembles.items():
                    MODEL_COUNT_COLUMNS.setdefault(name, f"{name}_count")

                    train_strict_acc, train_loose_acc, _, _ = compute_accuracy(ensemble_model, ensemble_train_loader, device)
                    val_strict_acc, val_loose_acc, y_true, y_pred = compute_accuracy(ensemble_model, ensemble_val_loader, device)

                    log_accuracy_to_csv(name, train_strict_acc, train_loose_acc, val_strict_acc, val_loose_acc)

                    results.append({
                        "Model Type": name,
                        "Train Strict Accuracy (%)": train_strict_acc,
                        "Train ±1 Grade Accuracy (%)": train_loose_acc,
                        "Val Strict Accuracy (%)": val_strict_acc,
                        "Val ±1 Grade Accuracy (%)": val_loose_acc,
                    })

                    # save_confusion_matrix_to_excel(y_true, y_pred, class_labels, name, excel_path)
                    export_predictions_to_excel(
                        ensemble_model,
                        ensemble_val_loader,
                        device,
                        grade_to_label,
                        excel_path,
                        sheet_name=f"{name}_preds",
                        model_name=name,
                    )

    df_results = pd.DataFrame(results)
    df_results_rounded = df_results.round(2)
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_results_rounded.to_excel(writer, sheet_name="Summary", index=False)
    print("=== Model Comparison Summary ===")
    print(df_results_rounded)
    return results


def summarize_accuracy_stability(
    all_results,
    excel_path="result/model_comparison_results.xlsx",
    sheet_name="Stability",
):
    """
    Aggregate per-model accuracy metrics across multiple runs and compute
    mean/std statistics to highlight training instability.
    """
    if not all_results:
        print("No accuracy records collected; skipping stability summary.")
        return pd.DataFrame()

    metric_keys = [
        ("Train Strict Accuracy (%)", "Train Strict"),
        ("Train ±1 Grade Accuracy (%)", "Train ±1 Grade"),
        ("Val Strict Accuracy (%)", "Val Strict"),
        ("Val ±1 Grade Accuracy (%)", "Val ±1 Grade"),
    ]
    aggregated = defaultdict(lambda: {key: [] for key, _ in metric_keys})
    for record in all_results:
        model_name = record.get("Model Type")
        if not model_name:
            continue
        target = aggregated[model_name]
        for key, _ in metric_keys:
            value = record.get(key)
            if value is not None:
                target[key].append(float(value))

    summary_rows = []
    for model_name, metric_lists in aggregated.items():
        row = {"Model Type": model_name}
        for key, label in metric_keys:
            values = metric_lists.get(key, [])
            if not values:
                continue
            row[f"{label} Mean (%)"] = float(np.mean(values))
            if len(values) > 1:
                row[f"{label} Std (%)"] = float(np.std(values, ddof=1))
            else:
                row[f"{label} Std (%)"] = 0.0
        summary_rows.append(row)

    if not summary_rows:
        print("Accuracy records contained no numeric values; skipping stability summary.")
        return pd.DataFrame()

    summary_df = pd.DataFrame(summary_rows).sort_values("Model Type")
    summary_df_rounded = summary_df.round(3)
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        summary_df_rounded.to_excel(writer, sheet_name=sheet_name, index=False)
    print("=== Accuracy Stability Summary ===")
    print(summary_df_rounded)
    return summary_df_rounded


def run_multiple_iterations(num_iterations=25, **compare_kwargs):
    """
    Execute `compare_models` multiple times and report aggregated accuracy statistics.
    """
    all_results = []
    for i in range(num_iterations):
        print(f"------------------------------------iteration no {i+1}------------------------------------")
        iteration_results = compare_models(**compare_kwargs)
        if iteration_results:
            all_results.extend(iteration_results)
    summarize_accuracy_stability(all_results)


# In[ ]:


# usage
if __name__ == "__main__":
    run_multiple_iterations(
        num_iterations = 25,
        model_types=list(BOOSTING_TYPES.keys()),
        include_ensemble=False
        )


# In[ ]:


# Ordinal evaluation helpers
def evaluate_ordinal_thresholds(model, loader, grade_to_label, device, decision_threshold=0.5, model_name=None, output_dir='./result'):
    model.eval()
    probs_list = []
    targets_list = []
    with torch.no_grad():
        for X, y in loader:
            inputs = tuple(x.to(device) for x in X)
            y = y.to(device)
            payload = inputs[0] if len(inputs) == 1 else inputs
            outputs = model(payload)
            if not isinstance(outputs, tuple):
                raise ValueError('Model is not configured for ordinal outputs.')
            probs, logits = outputs
            probs_list.append(probs.cpu())
            targets_list.append(y.cpu())
    if not probs_list:
        raise ValueError('No samples available for ordinal evaluation.')
    probs = torch.cat(probs_list, dim=0)
    targets = torch.cat(targets_list, dim=0)
    acc_per_threshold = threshold_accuracy(probs, targets, threshold=decision_threshold).cpu()
    grade_by_label = {v: k for k, v in grade_to_label.items()}
    threshold_labels = []
    for idx in range(acc_per_threshold.size(0)):
        grade = grade_by_label.get(idx, f'label_{idx}')
        threshold_labels.append(f"P(>{grade})")
    df = pd.DataFrame({
        'threshold': threshold_labels,
        'accuracy': (acc_per_threshold.numpy() * 100).round(2)
    })
    overall_pred = cumulative_to_labels(probs, threshold=decision_threshold)
    overall_acc = (overall_pred == targets).float().mean().item() * 100
    if model_name:
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f'ordinal_metrics_{model_name}.csv')
        df.to_csv(output_path, index=False)
        print(f'Saved threshold table to {output_path}')
    print(df)
    print(f'Overall accuracy: {overall_acc:.2f}%')
    return df, overall_acc



# In[ ]:


# --- Ordinal ensembles and multi-run utilities ---
ORDINAL_BASE_MODEL_TYPES = [
    'set_transformer_ordinal',
    'set_transformer_ordinal_xy',
    'set_transformer_ordinal_xy_additive',
    'deepset_ordinal',
    'deepset_ordinal_xy',
    'deepset_ordinal_xy_additive',
]

ORDINAL_ENSEMBLE_TYPES = [
    'ordinal_soft_voting_ensemble',
    'ordinal_stacking_ensemble',
    'ordinal_gbm_ensemble',
    'ordinal_xgboost_ensemble',
    'ordinal_adaboost_ensemble',
    # 'ordinal_lightgbm_ensemble',
    # 'ordinal_adaboost_ensemble',
    # 'ordinal_geometric_mean_ensemble',
    # 'ordinal_median_ensemble',
    # 'ordinal_trimmed_mean_ensemble',
]

def train_ordinal_stacking_meta_model(stacking_model, dataloader, device, epochs=5, lr=1e-3):
    if epochs <= 0:
        return
    stacking_model.meta_model.train()
    for member in stacking_model.models.values():
        member.eval()
    optimizer = torch.optim.Adam(stacking_model.meta_model.parameters(), lr=lr)
    for epoch in range(epochs):
        total_loss = 0.0
        total_samples = 0
        for X, y in dataloader:
            inputs = tuple(x.to(device) for x in X)
            targets = y.to(device)
            _, logits = stacking_model(inputs)
            loss = ordinal_logistic_loss(logits, targets)
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            batch_size = targets.size(0)
            total_loss += loss.item() * batch_size
            total_samples += batch_size
        if total_samples > 0:
            avg_loss = total_loss / total_samples
            print(f"Ordinal stacking meta epoch {epoch + 1}: loss={avg_loss:.4f}")
    stacking_model.meta_model.eval()


def infer_ordinal_stacking_feature_dim(stacking_model, dataloader, device):
    stacking_model.eval()
    with torch.no_grad():
        for X, _ in dataloader:
            inputs = tuple(x.to(device) for x in X)
            member_feats = stacking_model._member_features(inputs)
            M, _, F = member_feats.shape
            return F if stacking_model.combine == 'mean' else M * F
    raise RuntimeError('Unable to infer ordinal stacking feature dimension.')


def train_ordinal_tree_meta_model(tree_ensemble, dataloader, device):
    tree_ensemble.eval()
    for member in tree_ensemble.models.values():
        member.eval()
    feature_blocks = []
    target_blocks = []
    with torch.no_grad():
        for X, y in dataloader:
            inputs = tuple(x.to(device) for x in X)
            member_feats = tree_ensemble._member_features(inputs)
            feat = tree_ensemble._build_feature_matrix(member_feats)
            feature_blocks.append(feat.detach().cpu().numpy())
            target_blocks.append(y.detach().cpu().numpy())
    if not feature_blocks:
        raise RuntimeError('No data available to fit ordinal tree-based meta learner.')
    features = np.concatenate(feature_blocks, axis=0)
    targets = np.concatenate(target_blocks, axis=0)
    tree_ensemble.fit_meta_model(features, targets)


def build_ordinal_ensemble_models(
    ensemble_names,
    base_model_items,
    num_classes,
    device,
    train_loader,
    stacking_meta_epochs=5,
    stacking_meta_lr=1e-3,
    ensemble_weights=None,
    label_suffix='',
):
    ensembles = {}
    base_items = list(base_model_items)
    if not base_items:
        return ensembles

    def _resolve_group_weights(items):
        if ensemble_weights is None:
            return None
        if isinstance(ensemble_weights, dict):
            filtered = {name: ensemble_weights[name] for name, _ in items if name in ensemble_weights}
            if len(filtered) != len(items):
                missing = [name for name, _ in items if name not in filtered]
                if missing:
                    print(f"Warning: missing weights for {missing}; using uniform weights.")
                return None
            return filtered
        weight_list = list(ensemble_weights)
        if len(weight_list) != len(items):
            print('Warning: weight list length mismatch; using uniform weights.')
            return None
        return weight_list

    resolved_weights = _resolve_group_weights(base_items)

    for base_name in ensemble_names:
        cloned_items = [(name, copy.deepcopy(model)) for name, model in base_items]
        if not cloned_items:
            continue

        weights = resolved_weights
        if isinstance(weights, list):
            weights = list(weights)

        ensemble_key = f"{base_name}{label_suffix}" if label_suffix else base_name

        if base_name == 'ordinal_soft_voting_ensemble':
            ensemble_model = OrdinalSoftVotingEnsemble(cloned_items, weights=weights, freeze_members=True).to(device)
        elif base_name == 'ordinal_geometric_mean_ensemble':
            ensemble_model = OrdinalGeometricMeanEnsemble(cloned_items, weights=weights, freeze_members=True).to(device)
        elif base_name == 'ordinal_median_ensemble':
            ensemble_model = OrdinalMedianEnsemble(cloned_items, weights=weights, freeze_members=True).to(device)
        elif base_name == 'ordinal_trimmed_mean_ensemble':
            ensemble_model = OrdinalTrimmedMeanEnsemble(cloned_items, weights=weights, freeze_members=True, trim_frac=0.2).to(device)
        elif base_name == 'ordinal_stacking_ensemble':
            placeholder_dim = max(1, len(cloned_items) * (num_classes - 1))
            meta_model = nn.Linear(placeholder_dim, num_classes - 1).to(device)
            ensemble_model = OrdinalStackingEnsemble(
                cloned_items,
                num_classes=num_classes,
                weights=weights,
                freeze_members=True,
                meta_model=meta_model,
                feature_source='logits',
                combine='concat',
            ).to(device)
            inferred_dim = infer_ordinal_stacking_feature_dim(ensemble_model, train_loader, device)
            if inferred_dim != placeholder_dim:
                ensemble_model.meta_model = nn.Linear(inferred_dim, num_classes - 1).to(device)
            train_ordinal_stacking_meta_model(
                ensemble_model,
                train_loader,
                device,
                epochs=stacking_meta_epochs,
                lr=stacking_meta_lr,
            )
        elif base_name == 'ordinal_gbm_ensemble':
            try:
                ensemble_model = OrdinalGBMEnsemble(
                    cloned_items,
                    num_classes=num_classes,
                    weights=weights,
                    freeze_members=True,
                    feature_source='logits',
                    combine='concat',
                ).to(device)
            except ImportError as exc:
                print(f"Skipping {base_name}: {exc}")
                continue
            train_ordinal_tree_meta_model(ensemble_model, train_loader, device)
        elif base_name == 'ordinal_xgboost_ensemble':
            try:
                ensemble_model = OrdinalXGBoostEnsemble(
                    cloned_items,
                    num_classes=num_classes,
                    weights=weights,
                    freeze_members=True,
                    feature_source='logits',
                    combine='concat',
                ).to(device)
            except ImportError as exc:
                print(f"Skipping {base_name}: {exc}")
                continue
            train_ordinal_tree_meta_model(ensemble_model, train_loader, device)
        elif base_name == 'ordinal_lightgbm_ensemble':
            try:
                ensemble_model = OrdinalLightGBMEnsemble(
                    cloned_items,
                    num_classes=num_classes,
                    weights=weights,
                    freeze_members=True,
                    feature_source='logits',
                    combine='concat',
                ).to(device)
            except ImportError as exc:
                print(f"Skipping {base_name}: {exc}")
                continue
            train_ordinal_tree_meta_model(ensemble_model, train_loader, device)
        elif base_name == 'ordinal_adaboost_ensemble':
            ensemble_model = OrdinalAdaBoostEnsemble(cloned_items, weights=weights, freeze_members=True).to(device)
        else:
            print(f"Unknown ordinal ensemble type '{base_name}', skipping.")
            continue

        ensemble_model.eval()
        ensembles[ensemble_key] = ensemble_model

    return ensembles


def run_ordinal_iterations(
    ordinal_model_types=None,
    ordinal_ensemble_types=None,
    num_iterations=25,
    decision_threshold=0.5,
    stacking_meta_epochs=5,
    stacking_meta_lr=1e-3,
    ensemble_weights=None,
    output_excel='./result/ordinal_result.xlsx',
):
    ordinal_model_types = ordinal_model_types or ORDINAL_BASE_MODEL_TYPES
    ordinal_ensemble_types = ordinal_ensemble_types or ORDINAL_ENSEMBLE_TYPES
    threshold_records = []
    summary_records = []

    for iteration in range(num_iterations):
        print(f"----------------- Ordinal iteration {iteration + 1}/{num_iterations} -----------------")
        trained_base_models = {}
        base_dataset = None
        base_train_idx = None
        base_val_idx = None
        num_classes = None

        def record_metrics(model_name, table, overall_acc):
            for _, row in table.iterrows():
                threshold_records.append({
                    'iteration': iteration + 1,
                    'model': model_name,
                    'threshold': row['threshold'],
                    'accuracy': float(row['accuracy']),
                })
            summary_records.append({
                'iteration': iteration + 1,
                'model': model_name,
                'overall_accuracy': float(overall_acc),
            })

        for model_key in ordinal_model_types:
            print(f"=== Training ordinal model: {model_key} ===")
            train_loader, val_loader, model, dataset, train_idx, val_idx = main(model_key)
            trained_base_models[model_key] = model
            if base_dataset is None:
                base_dataset = dataset
                base_train_idx = train_idx
                base_val_idx = val_idx
            if num_classes is None:
                num_classes = getattr(model, 'num_classes', len(grade_to_label))

            table, overall_acc = evaluate_ordinal_thresholds(
                model,
                val_loader,
                grade_to_label=grade_to_label,
                device=device,
                decision_threshold=decision_threshold,
                model_name=None,
            )
            record_metrics(model_key, table, overall_acc)

        if ordinal_ensemble_types and trained_base_models:
            if base_dataset is None or base_train_idx is None or base_val_idx is None:
                raise RuntimeError('Dataset indices unavailable for ordinal ensembles.')
            collate_fn = make_collate_fn('set_transformer_xy')
            train_subset = Subset(base_dataset, base_train_idx)
            val_subset = Subset(base_dataset, base_val_idx)
            ensemble_train_loader = DataLoader(train_subset, batch_size=batch_size, shuffle=False, collate_fn=collate_fn)
            ensemble_val_loader = DataLoader(val_subset, batch_size=batch_size, shuffle=False, collate_fn=collate_fn)

            base_items = list(trained_base_models.items())
            ensemble_groups = [
                ("all", base_items),
                ("set_transformer", [(name, model) for name, model in base_items if "set_transformer" in name]),
                ("deepset", [(name, model) for name, model in base_items if "deepset" in name]),
            ]

            for group_name, group_items in ensemble_groups:
                if not group_items:
                    print(f"Skipping ordinal {group_name} ensembles (no models).")
                    continue

                ensembles = build_ordinal_ensemble_models(
                    ordinal_ensemble_types,
                    group_items,
                    num_classes=num_classes or len(grade_to_label),
                    device=device,
                    train_loader=ensemble_train_loader,
                    stacking_meta_epochs=stacking_meta_epochs,
                    stacking_meta_lr=stacking_meta_lr,
                    ensemble_weights=ensemble_weights,
                    label_suffix=f"_{group_name}",
                )

                for name, ensemble_model in ensembles.items():
                    table, overall_acc = evaluate_ordinal_thresholds(
                        ensemble_model,
                        ensemble_val_loader,
                        grade_to_label=grade_to_label,
                        device=device,
                        decision_threshold=decision_threshold,
                        model_name=None,
                    )
                    record_metrics(name, table, overall_acc)

    if not summary_records:
        print('No ordinal evaluations were executed.')
        return None

    threshold_df = pd.DataFrame(threshold_records)
    summary_df = pd.DataFrame(summary_records)

    threshold_agg = (threshold_df.groupby(['model', 'threshold'])['accuracy']
                     .agg(['mean', 'std'])
                     .reset_index()
                     .rename(columns={'mean': 'accuracy_mean', 'std': 'accuracy_std'}))
    summary_agg = (summary_df.groupby('model')['overall_accuracy']
                   .agg(['mean', 'std'])
                   .reset_index()
                   .rename(columns={'mean': 'overall_accuracy_mean', 'std': 'overall_accuracy_std'}))

    output_dir = os.path.dirname(output_excel)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        threshold_df.to_excel(writer, sheet_name='threshold_iterations', index=False)
        summary_df.to_excel(writer, sheet_name='overall_iterations', index=False)
        threshold_agg.to_excel(writer, sheet_name='threshold_avg', index=False)
        summary_agg.to_excel(writer, sheet_name='overall_avg', index=False)
    print(f"Saved aggregated ordinal results to {output_excel}")
    print(summary_agg.sort_values('model'))
    return {
        'threshold_iterations': threshold_df,
        'overall_iterations': summary_df,
        'threshold_summary': threshold_agg,
        'overall_summary': summary_agg,
    }


# In[ ]:


run_ordinal_iterations(num_iterations=25)

# run_ordinal_iterations(
#     ordinal_model_types=[
#         'set_transformer_ordinal',
#         'set_transformer_ordinal_xy',
#         'set_transformer_ordinal_xy_additive',
#         'deepset_ordinal',
#         'deepset_ordinal_xy',
#         'deepset_ordinal_xy_additive',
#     ],
#     ordinal_ensemble_types=[
#         'ordinal_soft_voting_ensemble',
#         'ordinal_geometric_mean_ensemble',
#         'ordinal_median_ensemble',
#         'ordinal_trimmed_mean_ensemble',
#         'ordinal_stacking_ensemble',
#         'ordinal_gbm_ensemble',
#         'ordinal_xgboost_ensemble',
#         'ordinal_lightgbm_ensemble',
#         'ordinal_adaboost_ensemble',
#     ],
#     num_iterations = num_iterations
# )
